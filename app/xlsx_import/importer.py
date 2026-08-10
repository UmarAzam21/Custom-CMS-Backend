import csv
import io
import re
import time
from dataclasses import dataclass
from typing import Callable, List, Optional

import openpyxl
import logging

from app.db import engine  # <-- adjust if your engine var is named differently

logger = logging.getLogger(__name__)


class ImportValidationError(Exception):
    """Raised when an import fails validation. Caller must roll back the new table."""


def _sanitize_column_name(raw: str, seen: set, idx: int) -> str:
    name = (raw or f"col_{idx}").strip().lower()
    name = re.sub(r"[^a-z0-9_]+", "_", name)
    name = re.sub(r"_+", "_", name).strip("_") or f"col_{idx}"
    if name[0].isdigit():
        name = f"c_{name}"
    original = name
    n = 1
    while name in seen:
        name = f"{original}_{n}"
        n += 1
    seen.add(name)
    return name


@dataclass
class ImportResult:
    table_name: str
    row_count: int
    columns: List[str]


def new_table_name() -> str:
    return f"dataset_{int(time.time() * 1000)}"


def read_header(filepath: str) -> List[str]:
    """Reads only the header row of the first sheet of a file - used as the
    canonical column set for the whole import, even across multiple files."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        if not wb.worksheets:
            raise ImportValidationError(f"{filepath}: workbook has no sheets.")
        ws = wb.worksheets[0]
        row_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration:
            raise ImportValidationError(f"{filepath}: first sheet is empty (no header row).")
        seen = set()
        columns = [
            _sanitize_column_name(str(h) if h is not None else "", seen, i)
            for i, h in enumerate(header_row)
        ]
        if not columns:
            raise ImportValidationError(f"{filepath}: no columns found.")
        return columns
    finally:
        wb.close()


def create_import_table(table_name: str, columns: List[str]):
    col_defs = ", ".join(f'"{c}" TEXT' for c in columns)
    tsvector_expr = " || ' ' || ".join(f"coalesce(\"{c}\", '')" for c in columns)
    conn = engine.raw_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            f'''
            CREATE TABLE "{table_name}" (
                id BIGSERIAL PRIMARY KEY,
                {col_defs},
                search_vector tsvector GENERATED ALWAYS AS (to_tsvector('english', {tsvector_expr})) STORED
            )
            '''
        )
        conn.commit()
        cur.close()
    finally:
        conn.close()


def _iter_all_rows(filepaths: List[str], columns: List[str]):
    """
    Yields data rows across EVERY sheet of EVERY file. Handles both common
    export shapes: header repeated on every sheet, or header only on the
    very first sheet of the very first file - either way it's skipped, only
    real data rows come out of this generator.
    """
    ncols = len(columns)
    for filepath in filepaths:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                first_row_of_sheet = True
                for row in ws.iter_rows(values_only=True):
                    if row is None:
                        continue
                    cells = list(row) + [None] * (ncols - len(row))
                    cells = ["" if v is None else str(v) for v in cells[:ncols]]
                    if all(c == "" for c in cells):
                        continue
                    if first_row_of_sheet:
                        first_row_of_sheet = False
                        seen = set()
                        as_header = [_sanitize_column_name(c, seen, i) for i, c in enumerate(cells)]
                        if as_header == columns:
                            continue  # this row was a repeated header - skip it
                    yield cells
        finally:
            wb.close()


def load_rows_into_table(
    table_name: str,
    columns: List[str],
    filepaths: List[str],
    batch_size: int = 20000,
    progress_cb: Optional[Callable[[int], None]] = None,
) -> int:
    """Streams rows from all files/sheets into the table via batched Postgres COPY."""
    conn = engine.raw_connection()
    total_rows = 0
    try:
        buf = io.StringIO()
        writer = csv.writer(buf)
        batched = 0

        def flush():
            nonlocal buf, writer, batched
            if batched == 0:
                return
            buf.seek(0)
            cur = conn.cursor()
            cols_sql = ", ".join(f'"{c}"' for c in columns)
            try:
                cur.copy_expert(f'COPY "{table_name}" ({cols_sql}) FROM STDIN WITH CSV', buf)
                conn.commit()
                logger.info("Flushed %d rows into %s", batched, table_name)
            except Exception:
                logger.exception("Error while flushing rows to %s", table_name)
                raise
            finally:
                cur.close()
                buf = io.StringIO()
                writer = csv.writer(buf)
                batched = 0

        logger.info("Starting import into %s from %d file(s)", table_name, len(filepaths))
        for cells in _iter_all_rows(filepaths, columns):
            writer.writerow(cells)
            total_rows += 1
            batched += 1
            if batched >= batch_size:
                flush()
                if progress_cb:
                    progress_cb(total_rows)
        flush()
        if progress_cb:
            progress_cb(total_rows)
        logger.info("Import finished into %s: %d total rows", table_name, total_rows)
        return total_rows
    except Exception:
        logger.exception("Import failed for %s, rolling back and dropping table", table_name)
        conn.rollback()
        cur = conn.cursor()
        cur.execute(f'DROP TABLE IF EXISTS "{table_name}"')
        conn.commit()
        cur.close()
        raise
    finally:
        conn.close()


def validate_import(result: ImportResult, min_rows: int = 1):
    """Extend with your real business rules - required columns, row-count deltas, etc."""
    if result.row_count < min_rows:
        raise ImportValidationError(
            f"Import produced {result.row_count} rows, expected at least {min_rows}."
        )
    if not result.columns:
        raise ImportValidationError("Import produced no columns.")


def build_indexes(table_name: str):
    conn = engine.raw_connection()
    try:
        cur = conn.cursor()
        cur.execute(f'CREATE INDEX ON "{table_name}" USING GIN (search_vector)')
        cur.execute(f'ANALYZE "{table_name}"')
        conn.commit()
        cur.close()
    finally:
        conn.close()


def activate_dataset(table_name: str):
    """
    Atomically points `current_dataset` at the new table, marks it active,
    retires the previous one, and drops its physical table. Blue/green swap.
    """
    conn = engine.raw_connection()
    try:
        cur = conn.cursor()
        cur.execute("SELECT table_name FROM dataset_versions WHERE status = 'active'")
        row = cur.fetchone()
        previous_table = row[0] if row else None

        cur.execute(f'CREATE OR REPLACE VIEW current_dataset AS SELECT * FROM "{table_name}"')
        cur.execute(
            "UPDATE dataset_versions SET status = 'active', activated_at = now() WHERE table_name = %s",
            (table_name,),
        )
        if previous_table:
            cur.execute(
                "UPDATE dataset_versions SET status = 'retired' WHERE table_name = %s",
                (previous_table,),
            )
        conn.commit()
        cur.close()

        if previous_table:
            cur = conn.cursor()
            cur.execute(f'DROP TABLE IF EXISTS "{previous_table}"')
            conn.commit()
            cur.close()
    finally:
        conn.close()